import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill
import os
import shutil

# --- CẤU HÌNH ---
# !!! Vui lòng thay đổi các giá trị dưới đây cho phù hợp !!!

# 1. Đường dẫn đến file Excel chứa thông tin âm thanh
EXCEL_FILE_PATH = r'F:\Toàn-Khang_NCKH\ngtai_dataset\The UK Covid-19\metadatabinary.xlsx'

# 2. Đường dẫn đến file Excel ĐẦU RA (đã được tô màu)
#    Script sẽ không ghi đè lên file gốc của bạn.
OUTPUT_EXCEL_PATH = r'F:\Toàn-Khang_NCKH\ngtai_dataset\The UK Covid-19\metadatabinary_PROCESSED.xlsx'

# 3. Tên thư mục chứa các file âm thanh GỐC
SOURCE_AUDIO_FOLDER = r'F:\Toàn-Khang_NCKH\ngtai_dataset\The UK Covid-19 zip\audio'

# 4. Tên thư mục ĐÍCH chính để chứa các thư mục con
DESTINATION_FOLDER = r'F:\Toàn-Khang_NCKH\ngtai_dataset\datasetbinary'

# ---------------------------------------------------------

def process_audio_files_and_highlight(excel_path, output_excel_path, source_folder, dest_folder):
    """
    Đọc file Excel, tạo thư mục con, sao chép/đổi tên file, và tô màu
    các ô đã xử lý thành công.
    """
    # Kiểm tra sự tồn tại của file Excel
    try:
        wb = openpyxl.load_workbook(excel_path)
        print(f"Đã mở file Excel: '{excel_path}'")
    except FileNotFoundError:
        print(f"Lỗi: Không tìm thấy file Excel tại '{excel_path}'.")
        return

    # Định nghĩa màu để tô (xanh lá cây nhạt)
    highlight_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    
    # Tạo thư mục đích chính nếu chưa có
    os.makedirs(dest_folder, exist_ok=True)

    sheet_names = ['asthma', 'covid', 'healthy']
    total_files_copied = 0
    total_files_not_found = 0

    # Lặp qua từng sheet
    for sheet_name in sheet_names:
        if sheet_name not in wb.sheetnames:
            print(f"Cảnh báo: Không tìm thấy sheet '{sheet_name}' trong file Excel. Bỏ qua.")
            continue
            
        ws = wb[sheet_name]
        print(f"\n--- Đang xử lý sheet: '{sheet_name}' ---")

        # Tạo thư mục con tương ứng với tên sheet
        sheet_specific_folder = os.path.join(dest_folder, sheet_name)
        os.makedirs(sheet_specific_folder, exist_ok=True)

        # Tìm chỉ số cột (column index) dựa trên tên tiêu đề
        header = [cell.value for cell in ws[1]]
        try:
            id_col_idx = header.index('participant_identifier')
            cough_col_idx = header.index('cough_file_name')
            three_cough_col_idx = header.index('three_cough_file_name')
        except ValueError as e:
            print(f"Lỗi: Thiếu cột trong sheet '{sheet_name}'. Chi tiết: {e}")
            continue

        # Lặp qua từng hàng (bắt đầu từ hàng 2 để bỏ qua tiêu đề)
        for row in ws.iter_rows(min_row=2):
            participant_id = row[id_col_idx].value
            
            # Xử lý 2 cột filename
            file_cells_to_check = [row[cough_col_idx], row[three_cough_col_idx]]

            for file_cell in file_cells_to_check:
                original_filename = file_cell.value
                
                if not original_filename or not isinstance(original_filename, str):
                    continue

                source_file_path = os.path.join(source_folder, original_filename)

                if os.path.exists(source_file_path):
                    new_filename = f"{participant_id}_{original_filename}"
                    destination_file_path = os.path.join(sheet_specific_folder, new_filename)
                    
                    shutil.copy2(source_file_path, destination_file_path)
                    
                    # *** TÍNH NĂNG MỚI: TÔ MÀU Ô ĐÃ XỬ LÝ ***
                    file_cell.fill = highlight_fill

                    print(f"  [OK] Đã xử lý '{original_filename}' và tô màu ô.")
                    total_files_copied += 1
                else:
                    print(f"  [LỖI] Không tìm thấy file '{original_filename}' trong '{source_folder}'")
                    total_files_not_found += 1
    
    # Lưu các thay đổi (tô màu) vào một file Excel MỚI
    try:
        wb.save(output_excel_path)
        print(f"\n--- HOÀN TẤT ---")
        print(f"Tổng số file đã sao chép: {total_files_copied}")
        print(f"Tổng số file không tìm thấy: {total_files_not_found}")
        print(f"File Excel đã tô màu được lưu tại: '{output_excel_path}'")
        print(f"Các file âm thanh đã xử lý được lưu trong '{dest_folder}'")
    except Exception as e:
        print(f"Lỗi khi lưu file Excel: {e}")


# Chạy hàm chính
if __name__ == "__main__":
    process_audio_files_and_highlight(
        EXCEL_FILE_PATH,
        OUTPUT_EXCEL_PATH,
        SOURCE_AUDIO_FOLDER,
        DESTINATION_FOLDER
    )