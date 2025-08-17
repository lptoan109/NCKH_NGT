import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill

def highlight_matching_ids(disease_file, sound_file, output_file):
    """
    Đọc ID từ nhiều sheet trong file bệnh và tô màu các ID trùng khớp
    trong file âm thanh.

    Args:
        disease_file (str): Đường dẫn đến file Excel thông tin bệnh.
        sound_file (str): Đường dẫn đến file Excel thông tin âm thanh.
        output_file (str): Đường dẫn để lưu file kết quả đã được tô màu.
    """
    # 1. Định nghĩa các sheet và màu sắc tương ứng (mã màu HEX)
    # Bạn có thể thay đổi mã màu tại đây
    sheets_and_colors = {
        'covid': 'FFC7CE',  # Màu đỏ nhạt
        'healthy': 'C6EFCE',  # Màu xanh lá nhạt
        'asthma': 'BDD7EE'    # Màu xanh dương nhạt
    }

    print("Bắt đầu đọc ID từ file bệnh...")
    
    # 2. Đọc tất cả ID từ các sheet được chỉ định trong file bệnh
    # Sử dụng set để tối ưu hóa tốc độ tìm kiếm (O(1))
    all_disease_ids = {}
    for sheet_name in sheets_and_colors.keys():
        try:
            df_disease = pd.read_excel(disease_file, sheet_name=sheet_name)
            # Loại bỏ các giá trị rỗng và chuyển thành một set
            all_disease_ids[sheet_name] = set(df_disease['participant_identifier'].dropna())
            print(f"Đã đọc thành công {len(all_disease_ids[sheet_name])} ID từ sheet '{sheet_name}'.")
        except FileNotFoundError:
            print(f"Lỗi: Không tìm thấy file bệnh tại '{disease_file}'. Vui lòng kiểm tra lại.")
            return
        except Exception as e:
            print(f"Lỗi khi đọc sheet '{sheet_name}': {e}. Bỏ qua sheet này.")
            all_disease_ids[sheet_name] = set()

    # 3. Mở file âm thanh bằng openpyxl để tiến hành tô màu
    try:
        wb = openpyxl.load_workbook(sound_file)
        ws = wb.active  # Giả sử làm việc trên sheet đầu tiên (active sheet)
        print(f"\nĐã mở file âm thanh: '{sound_file}' và làm việc trên sheet '{ws.title}'.")
    except FileNotFoundError:
        print(f"Lỗi: Không tìm thấy file âm thanh tại '{sound_file}'. Vui lòng kiểm tra lại đường dẫn.")
        return

    # 4. Tìm vị trí cột 'participant_identifier' trong file âm thanh
    header = [cell.value for cell in ws[1]]
    try:
        # Cột trong openpyxl bắt đầu từ 1
        id_column_index = header.index('participant_identifier') + 1
    except ValueError:
        print(f"Lỗi: Không tìm thấy cột 'participant_identifier' trong file âm thanh.")
        return

    print("Bắt đầu quét và tô màu các ID trùng khớp...")
    
    # 5. Lặp qua từng dòng trong file âm thanh để kiểm tra và tô màu
    highlight_count = 0
    # Bắt đầu từ hàng 2 để bỏ qua tiêu đề
    for row_index in range(2, ws.max_row + 1):
        cell = ws.cell(row=row_index, column=id_column_index)
        participant_id = cell.value

        if participant_id:
            # Kiểm tra ID có trong danh sách của sheet nào không
            for sheet_name, ids_set in all_disease_ids.items():
                if participant_id in ids_set:
                    # Lấy màu từ dictionary và tạo fill
                    color_hex = sheets_and_colors[sheet_name]
                    fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")
                    # Áp dụng màu cho ô chứa ID
                    cell.fill = fill
                    highlight_count += 1
                    # Dừng tìm kiếm khi đã tìm thấy và tô màu
                    break

    # 6. Lưu file kết quả
    try:
        wb.save(output_file)
        print(f"\nHoàn thành! Đã tô màu {highlight_count} ô.")
        print(f"File kết quả đã được lưu tại: '{output_file}'")
    except Exception as e:
        print(f"Lỗi khi lưu file: {e}")

# --- Cấu hình tên file ---
# !!! QUAN TRỌNG: Hãy thay đổi tên file dưới đây cho phù hợp với bạn !!!
DISEASE_EXCEL_FILE = r'H:\Toàn-Khang_NCKH\ngtai_dataset\The UK Covid-19\participant_metadata.xlsx'
SOUND_EXCEL_FILE = r'H:\Toàn-Khang_NCKH\ngtai_dataset\The UK Covid-19\audio_metadata.xlsx'
OUTPUT_EXCEL_FILE = r'H:\Toàn-Khang_NCKH\ngtai_dataset\The UK Covid-19\audio_metadata_highlighted.xlsx'

# --- Chạy hàm chính ---
if __name__ == "__main__":
    highlight_matching_ids(DISEASE_EXCEL_FILE, SOUND_EXCEL_FILE, OUTPUT_EXCEL_FILE)